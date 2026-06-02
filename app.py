import streamlit as st
import pandas as pd
import re
from rapidfuzz import process
from datetime import datetime
import io
from openpyxl import load_workbook

# ============================================================
# CRM FIELD LIST
# ============================================================
CRM_FIELDS = [
    'Campaign', 'CustomerType', 'customerName', 'CustomerSubtype',
    'ContactNumber', 'City', 'Location', 'SubLocation', 'Area',
    'Address', 'Email', 'Facilities', 'ReferenceId', 'CustomerId',
    'ClientId', 'CustomerDate', 'CustomerYear', 'Other', 'Description',
    'Video', 'GoogleMap', 'Price', 'LeadType', 'URL', 'isFavourite', 'Verified'
]

# ============================================================
# JAIPUR KE POPULAR AREAS KI LIST
# ============================================================
# JAIPUR_AREAS = [
#     'Vaishali Nagar', 'Mansarovar', 'Jagatpura', 'Malviya Nagar', 'Tonk Road',
#     'Ajmer Road', 'Sanganer', 'Bani Park', 'C Scheme', 'Raja Park', 'Jhotwara',
#     'Kalwar Road', 'Pratap Nagar', 'Gopalpura', 'Sikar Road', 'Sitapura',
#     'Vidhyadhar Nagar', 'VKI', 'Shyam Nagar', 'Sodala', 'Civil Lines',
#     'Lalkothi', 'Durgapura', 'Mahesh Nagar', 'Patrakar Colony', 'Nirman Nagar',
#     'Muhana', 'Sirsi Road', 'Bindayaka', 'Kanakpura', 'Chitrakoot',
#     'Heerapura', 'Murlipura', 'Hasanpura', 'Adarsh Nagar', 'Jawahar Nagar',
#     'Transport Nagar', 'Bapu Nagar', 'Tilak Nagar', 'Triveni Nagar',
#     'New Sanganer Road', 'Ramnagariya', 'Airport Road', 'Narayan Vihar',
#     'Jagdamba Nagar', 'Khatipura', 'Brahmpuri', 'Govindpura', 'Ambabari',
#     'Nehru Nagar', 'Lal Kothi', 'JLN Marg', 'Ashok Nagar', 'Shastri Nagar',
#     'Banipark', 'Bais Godam', 'Jawahar Circle', 'Gandhi Path',
#     'Rohini Nagar', 'Patel Nagar', 'Hari Marg', 'Sector 1', 'Sector 2',
#     'Sector 3', 'Sector 4', 'Sector 5', 'Sector 7', 'Sector 9'
# ]

JAIPUR_LOCATION_MASTER = {

    "Jagatpura": [
        "jagatpura", "new jagatpura", "model town jagatpura",
    "ramnagariya", "mahal road", "vit road", "skit road",
    "akshay patra road", "gyan vihar", "shivam nagar",
    "nri colony", "rseb colony", "ashadeep green avenue",
    "ashadeep rainbow", "vrinda gardens", "dmart jagatpura",
    "rukmani vihar", "kalyanpura", "tejaji nagar",
    "shiv shakti nagar", "chokhi dhani road", "kumbha marg",
    "kusum vihar", "green park jagatpura", "apex circle",
    "goner road", "chaksu road"
    ],


    "Mansarovar": [
        "Mansarovar", "Patrakar Colony", "Agarwal Farm",
        "VT Road", "Shipra Path", "Rajat Path",
        "Madhyam Marg", "Varun Path",
        "Kiran Path", "New Sanganer Road",
        "Ricco Mansarovar", "City Park",
        "Mansarovar Metro", "Thadi Market",
        "Heera Path", "Meera Marg"
    ],

    "Malviya Nagar": [
        "Malviya Nagar", "Jawahar Circle",
        "JLN Marg", "Airport Road",
        "Bajaj Nagar", "Calgary Road",
        "Gaurav Tower", "GT",
        "World Trade Park", "WTP",
        "Fortis Hospital", "Apex Circle",
        "D Block Malviya Nagar",
        "Satkar Shopping Center"
    ],

    "Vaishali Nagar": [
        "Vaishali Nagar", "Gandhi Path",
        "Queens Road", "Nemi Nagar",
        "Hanuman Nagar", "Chitrakoot",
        "Amrapali Circle", "Nursery Circle",
        "Sirsi Road", "Kanakpura",
        "National Handloom", "Elements Mall",
        "Vaibhav Tower", "Prince Road"
    ],

    "Pratap Nagar": [
        "Pratap Nagar", "Sector 3", "Sector 5",
        "Sector 7", "Sector 8", "Sector 11",
        "Sector 17", "Sector 21",
        "Kumbha Marg", "NRI Circle",
        "Airport Terminal 2",
        "Sanganer Airport"
    ],

    "Sitapura": [
        "Sitapura", "EPIP",
        "Sitapura Industrial Area",
        "JECRC University",
        "Chokhi Dhani",
        "Mahindra SEZ",
        "India Gate Sitapura",
        "Genpact", "Infosys",
        "Mahindra World City"
    ],

    "Sanganer": [
        "Sanganer", "Muhana",
        "Muhana Mandi", "Diggi Road",
        "Vatika Road", "Ramchandrapura",
        "Shivdaspura", "Tonk Phatak",
        "Kalyanpura", "Sanganer Bazar"
    ],

    "Vidhyadhar Nagar": [
        "Vidhyadhar Nagar",
        "Central Spine",
        "Sector 1", "Sector 2",
        "Sector 3", "Sector 4",
        "Sector 5", "Sector 7",
        "Sector 9", "Alka Cinema",
        "Temple Road VDN"
    ],

    "Jhotwara": [
        "Jhotwara", "Kalwar Road",
        "Govindpura", "Nivaru Road",
        "Murlipura", "Benad Road",
        "Triton Mall", "VKI Road",
        "Khatipura Road"
    ],

    "Ajmer Road": [
        "Ajmer Road", "Bhankrota",
        "Heerapura", "Mahapura",
        "Omaxe City", "Kamla Nehru Nagar",
        "200 Feet Bypass",
        "Pink Pearl", "RIICO Bhankrota"
    ],

    "Tonk Road": [
        "Tonk Road", "Lalkothi",
        "Durgapura", "Gopalpura",
        "Airport Circle",
        "Gandhi Nagar",
        "Tonk Phatak",
        "Sawai Madhopur Road"
    ],

    "Civil Lines": [
        "Civil Lines", "Hathroi",
        "Shyam Nagar", "Sodala",
        "Bais Godam", "Hasanpura",
        "Jyoti Nagar",
        "High Court Circle"
    ],

    "Raja Park": [
        "Raja Park", "Adarsh Nagar",
        "Tilak Nagar", "Jawahar Nagar",
        "Bapu Nagar", "Transport Nagar",
        "Pink Square Mall",
        "Govind Marg"
    ]
}




# ============================================================
# FUNCTIONS
# ============================================================


def extract_area_and_bhk(text):

    if not isinstance(text, str):
        return "N/A"

    # ------------------------------------------------
    # SQFT PRIORITY
    # ------------------------------------------------
    sqft_pattern = r'\d+\s*(?:sq\s*ft|sqft|square\s*feet)'

    sqft_match = re.search(sqft_pattern, text, re.IGNORECASE)

    if sqft_match:
        return sqft_match.group().strip()

    # ------------------------------------------------
    # SQ YARD PRIORITY
    # ------------------------------------------------
    yard_pattern = r'\d+\s*(?:sq\s*yard|sq\s*yd|sq\s*yards)'

    yard_match = re.search(yard_pattern, text, re.IGNORECASE)

    if yard_match:
        return yard_match.group().strip()

    # ------------------------------------------------
    # BHK FALLBACK
    # ------------------------------------------------
    bhk_pattern = r'\b\d+\s*BHK\b'

    bhk_match = re.search(bhk_pattern, text, re.IGNORECASE)

    if bhk_match:
        return bhk_match.group().strip()

    return "N/A"



def extract_clean_prices(text):

    if not isinstance(text, str):
        return "N/A"

    text = text.replace(",", "")
    lower_text = text.lower()

    # ------------------------------------------------
    # MAIN PRICE KEYWORDS ONLY
    # ------------------------------------------------
    patterns = [

        # Rent 25000
        r'(?:rent|price|budget|cost)\D{0,15}(₹?\s*\d+(?:\.\d+)?\s*(?:k|thousand|lakh|lac|crore|cr)?)',

        # ₹25000
        r'₹\s*(\d+(?:\.\d+)?)',

        # Rs 25000
        r'rs\.?\s*(\d+(?:\.\d+)?)'
    ]

    candidate = None

    for pattern in patterns:

        match = re.search(pattern, lower_text, re.IGNORECASE)

        if match:

            candidate = match.group(1).strip()
            break

    if not candidate:
        return "N/A"

    # ------------------------------------------------
    # REMOVE PHONE NUMBERS
    # ------------------------------------------------
    digits_only = re.sub(r'\D', '', candidate)

    if len(digits_only) >= 10:
        return "N/A"

    # ------------------------------------------------
    # NORMALIZATION
    # ------------------------------------------------
    candidate = candidate.replace("₹", "")
    candidate = candidate.replace("rs", "")
    candidate = candidate.replace(".", "")
    candidate = candidate.strip()

    # thousand → k
    thousand_match = re.search(r'(\d+(?:\.\d+)?)\s*thousand', candidate, re.IGNORECASE)

    if thousand_match:

        num = thousand_match.group(1)
        return f"{num}k"

    # lakh → Lac
    candidate = re.sub(r'lakh', 'Lac', candidate, flags=re.IGNORECASE)

    # crore → Cr
    candidate = re.sub(r'crore', 'Cr', candidate, flags=re.IGNORECASE)


    candidate = candidate.strip().lower()



    if re.fullmatch(r'\d+(?:\.\d+)?\s*k', candidate):

        num_match = re.search(r'\d+(?:\.\d+)?', candidate)

        if num_match:
            return f"{num_match.group()}k"

    # ------------------------------------------------
    # PLAIN NUMBER → K
    # ------------------------------------------------
    plain_match = re.fullmatch(r'\d+', candidate)

    if plain_match:

        num = int(candidate)

        # IGNORE TOO SMALL
        if num < 1000:
            return "N/A"

        # 25000 → 25k
        if num >= 1000 and num < 1000000:
            return f"{round(num / 1000)}k"

    # ------------------------------------------------
    # FINAL CLEAN
    # ------------------------------------------------
    candidate = candidate.replace(" ", "")

    return candidate
    return candidate.strip()

from rapidfuzz import fuzz

def extract_geo_data(text):


    text_str = str(text)

    found_loc = "N/A"
    found_sub_loc = "N/A"
    found_address = "N/A"

    # CLEAN TEXT
    clean_text = re.sub(r'[^a-zA-Z0-9\s]', ' ', text_str)
    clean_text = re.sub(r'\s+', ' ', clean_text).strip()

    # ------------------------------------------------
    # STEP 1 : EXACT MATCH
    # ------------------------------------------------
    # for area in JAIPUR_AREAS:

    # for main_area, sub_locations in JAIPUR_LOCATION_MASTER.items():
    #
    #     for area in sub_locations:
    #
    #         if area.lower() in clean_text.lower():
    #             found_loc = main_area
    #             found_sub_loc = area
    #             found_address = f"{area}, Jaipur"
    #
    #             return found_loc, found_sub_loc, found_address
    #
    #     if area.lower() in clean_text.lower():
    #
    #         found_loc = area
    #         found_sub_loc = area
    #         found_address = f"{area}, Jaipur"
    #
    #         return found_loc, found_sub_loc, found_address
    # ------------------------------------------------
    # STEP 1 : EXACT MATCH
    # ------------------------------------------------

    for main_area, sub_locations in JAIPUR_LOCATION_MASTER.items():

        for area in sub_locations:

            if area.lower() in clean_text.lower():
                found_loc = main_area
                found_sub_loc = area
                found_address = f"{area}, Jaipur"

                return found_loc, found_sub_loc, found_address

    # ------------------------------------------------
    # STEP 2 : FUZZY MATCH
    # ------------------------------------------------
    words = clean_text.split()

    # for area in JAIPUR_LOCATION_MASTER:
    #
    #     area_lower = area.lower()
    #
    #     for i in range(len(words)):
    #
    #         chunk = " ".join(words[i:i+3]).lower()
    #
    #         score = fuzz.ratio(chunk, area_lower)
    #
    #         if score >= 80:
    #
    #             found_loc = area
    #             found_sub_loc = area
    #             found_address = f"{area}, Jaipur"
    #
    #             return found_loc, found_sub_loc, found_address

    for main_area, sub_locations in JAIPUR_LOCATION_MASTER.items():

        for area in sub_locations:

            area_lower = area.lower()

            for i in range(len(words)):

                chunk = " ".join(words[i:i + 3]).lower()

                score = fuzz.ratio(chunk, area_lower)

                if score >= 80:
                    found_loc = main_area
                    found_sub_loc = area
                    found_address = f"{area}, Jaipur"

                    return found_loc, found_sub_loc, found_address

    return found_loc, found_sub_loc, found_address


def extract_indian_mobile(text):
    if pd.isna(text) or str(text).strip() == "": return "1010101010"
    clean_text = re.sub(r'[\s\-\(\)\.\/]', '', str(text))
    matches = re.findall(r'(?:(?<=^)|(?<=\D))([6-9]\d{9})(?=\D|$)', clean_text)
    if not matches:
        twelve_digit = re.findall(r'91([6-9]\d{9})', clean_text)
        if twelve_digit:
            matches = twelve_digit
    return matches[0] if matches else "1010101010"

def get_customer_type_and_subtype(text):
    t = text.lower()
    c_type, c_subtype = "Other", "N/A"
    bhk_num_match = re.search(r'(\d+)\s*bhk', t)
    if bhk_num_match or 'flat' in t or 'apartment' in t:
        c_type = "Flat"
        if bhk_num_match:
            num = int(bhk_num_match.group(1))
            if num == 1: c_subtype = "1 BHK"
            elif num == 2: c_subtype = "2 BHK"
            elif num >= 3: c_subtype = "3 BHK and above"
        elif any(x in t for x in ['3+1', '3+study', '4 bhk', '5 bhk']):
            c_subtype = "3 BHK and above"
    elif any(x in t for x in ['house', 'independent makaan', 'duplex', 'makaan']):
        c_type, c_subtype = "House", "Independent Makaan/Duplex"
    elif any(x in t for x in ['villa', 'bungalow', 'kothi']):
        c_type, c_subtype = "Villa", "Villa/Bungalow"
    elif any(x in t for x in ['room', 'single room', 'portion']):
        c_type, c_subtype = "Room", "Single Room/Portion"
    elif any(x in t for x in ['shop', 'office', 'warehouse', 'showroom', 'basement', 'commercial']):
        c_type, c_subtype = "Commercial Space", "Shop/Office/Commercial"
    elif any(x in t for x in ['plot', 'jda approved']):
        c_type, c_subtype = "Residential Plot", "Plot/JDA Plot"
    elif 'farm house' in t:
        c_type, c_subtype = "Farm House", "Farm House"
    elif any(x in t for x in ['bigha', 'kheti', 'agricultural']):
        c_type, c_subtype = "Agricultural Land", "Bigha/Kheti Land"
    return c_type, c_subtype

def clean_description(text):

    if not isinstance(text, str):
        return "N/A"

    # -----------------------------------
    # REMOVE HASHTAGS
    # -----------------------------------
    text = re.sub(r'#\w+', ' ', text)

    # -----------------------------------
    # REMOVE EMOJIS & SPECIAL SYMBOLS
    # -----------------------------------

    # REMOVE EMOJIS ONLY (Hindi safe)
    text = re.sub(
        r'[\U00010000-\U0010ffff]',
        ' ',
        text
    )
    # KEEP HINDI + ENGLISH + NUMBERS
    text = re.sub(r'[^a-zA-Z0-9\u0900-\u097F\s₹.,:/&()%-]', ' ', text)

    # -----------------------------------
    # REMOVE EXTRA COMMAS / DOTS
    # -----------------------------------
    text = re.sub(r',+', ' ', text)
    text = re.sub(r'\.+', '.', text)

    # -----------------------------------
    # REMOVE USELESS WORDS
    # -----------------------------------
    useless_words = [
        'luxuryhomes',
        'luxurymansion',
        'modernhomes',
        'milliondollarlisting',
        'realtor',
        'realtorlife',
        'realestateinfluencer',
        'homedecor',
        'interiordesign',
        'beautiful',
        'dreamhome',
        'hometour',
        'dm for more information',
        'dm for more info',
        'site visit',
        'call now',
        'creating legacy',
        'building trust'
    ]

    marketing_lines = [

        'jeevan shaili',
        'lifestyle',
        'dream life',
        'dream living',
        'premium living',
        'luxury living',
        'township ka matlab',
        'sirf plotting nahi',
        'sirf plots nahi',
        'poori jeevan shaili',
        'perfect lifestyle',
        'best investment',
        'golden opportunity',
        'future investment',
        'investment opportunity',
        'live the luxury',
        'live your dream',
        'modern lifestyle',
        'high class living',
        'elite living',
        'luxurious lifestyle',
        'world class lifestyle'
    ]

    for word in useless_words:
        text = re.sub(word, ' ', text, flags=re.IGNORECASE)

    # -----------------------------------
    # KEEP ONLY IMPORTANT LINES
    # -----------------------------------
    important_keywords = [
        'bhk',
        'flat',
        'plot',
        'villa',
        'house',
        'farmhouse',
        'sq ft',
        'sq yd',
        'sq yds',
        'loan',
        'price',
        'rate',
        'jaipur',
        'jda',
        'rera',
        'club house',
        'gym',
        'swimming',
        'location',
        'project',
        'commercial',
        'office',
        'warehouse',
        'shop',
        'balcony',
        'security',
        'lift',
        'garden',
        'road',
        'parking',
        'cctv'
    ]

    final_lines = []



    # for line in text.split("\n"):
    #
    #     clean_line = line.strip()
    #
    #     if len(clean_line) < 5:
    #         continue
    #
    #     # REMOVE MARKETING LINES
    #     if any(word.lower() in clean_line.lower() for word in marketing_lines):
    #         continue
    #
    #     if any(keyword.lower() in clean_line.lower() for keyword in important_keywords):
    #         final_lines.append(clean_line)

    for line in text.split("\n"):

        clean_line = line.strip()

        if len(clean_line) < 5:
            continue

        # NORMALIZE LINE
        normalized_line = re.sub(
            r'[^a-zA-Z0-9\u0900-\u097F\s]',
            '',
            clean_line.lower()
        )

        normalized_line = re.sub(r'\s+', ' ', normalized_line).strip()

        # REMOVE MARKETING LINES
        if any(
                re.sub(r'\s+', ' ', word.lower()).strip() in normalized_line
                for word in marketing_lines
        ):
            continue

        if any(keyword.lower() in clean_line.lower() for keyword in important_keywords):
            final_lines.append(clean_line)

    # -----------------------------------
    # FINAL CLEANING
    # -----------------------------------
    final_text = " | ".join(final_lines)

    final_text = re.sub(r'\s+', ' ', final_text).strip()

    if final_text == "":
        return "N/A"

    return final_text

# ============================================================
# STREAMLIT UI
# ============================================================
st.set_page_config(page_title="Real Estate Miner", page_icon="🏢")
st.title("🏢 Real Estate Data Miner")
st.info("UPLOAD EXCEL FILE")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])
selected_area = st.selectbox(
    "Select Jaipur Area",
    ["ALL JAIPUR"] + list(JAIPUR_LOCATION_MASTER.keys())
)

# ============================================================
# MAIN PROCESSING
# ============================================================
if uploaded_file:
    try:
        df_raw = pd.read_excel(uploaded_file, sheet_name='Data')
        st.success(f"File Loading! Total {len(df_raw)} records.")

        if st.button("Start Mining 🚀"):
            current_date = datetime.now().strftime('%d-%m-%Y')
            rows = []
            bar = st.progress(0)

            # Check columns to decide source
            has_caption = 'caption' in df_raw.columns
            has_text = 'text' in df_raw.columns

            for i, (_, row) in enumerate(df_raw.iterrows()):
                # Logic to select source text and ReferenceId
                if has_caption:
                    raw_text = str(row.get('caption', "")).strip()
                    ref_id = 'instagram'
                elif has_text:
                    raw_text = str(row.get('text', "")).strip()
                    ref_id = 'facebook'
                else:
                    st.error("Excel mein na 'text' column mila na 'caption'! Mining rok di gayi.")
                    st.stop()

                if raw_text == "" or raw_text.lower() == 'nan':
                    continue

                entry = {field: "N/A" for field in CRM_FIELDS}

                entry.update({
                    'ContactNumber': str(extract_indian_mobile(raw_text)),
                    'Description': clean_description(raw_text),
                    'Area': extract_area_and_bhk(raw_text),
                    'Price': extract_clean_prices(raw_text),
                    'City': 'Jaipur',
                    'CustomerDate': str(current_date),
                    'ReferenceId': ref_id  # Automatic Switch
                })

                c_type, c_subtype = get_customer_type_and_subtype(raw_text)
                entry['CustomerType'] = c_type
                entry['CustomerSubtype'] = c_subtype

                # loc, sub_loc, addr = extract_geo_data(raw_text)
                # if selected_area != "ALL JAIPUR":
                #
                #     if loc != selected_area:
                #         continue
                # entry['Location'] = loc
                # entry['SubLocation'] = sub_loc
                # entry['Address'] = addr

                loc, sub_loc, addr = extract_geo_data(raw_text)

                # Jaipur ke bahar ka data ignore
                if loc == "N/A":
                    continue

                # Specific area filter
                if selected_area != "ALL JAIPUR":
                    if loc != selected_area:
                        continue

                entry['Location'] = loc
                entry['SubLocation'] = sub_loc
                entry['Address'] = addr

                lower_text = raw_text.lower()
                if 'rent' in lower_text:
                    entry['Campaign'] = 'Rentout'
                elif any(x in lower_text for x in ['sale', 'sell']):
                    entry['Campaign'] = 'Seller'
                elif any(x in lower_text for x in ['buy', 'looking for']):
                    entry['Campaign'] = 'Buyer'

                rows.append(entry)
                bar.progress((i + 1) / len(df_raw))

            df_final = pd.DataFrame(rows)
            st.write(f"### Data Preview (Detected Source: {ref_id.capitalize()}):")
            st.dataframe(df_final.head(10))

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name='ProcessedData')
                workbook = writer.book
                worksheet = writer.sheets['ProcessedData']
                headers = [cell.value for cell in worksheet[1]]

                # Number format fix for Excel
                for col_name in ['ContactNumber', 'CustomerDate']:
                    if col_name in headers:
                        col_idx = headers.index(col_name) + 1
                        for row_cells in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                            for cell in row_cells: cell.number_format = '@'

            output.seek(0)
            st.download_button(
                label="Download Processed Excel 📥",
                data=output,
                file_name=f"Processed_{ref_id}_{current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Error: {e}")
